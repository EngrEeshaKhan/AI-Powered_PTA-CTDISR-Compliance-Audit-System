from pathlib import Path

from openpyxl import load_workbook


def parse_excel(file_path: str | Path) -> dict:
    """
    Extract structured information from an XLSX workbook.
    """

    path = Path(file_path)

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
    )

    sheets = []
    total_rows = 0

    for worksheet in workbook.worksheets:

        rows = []

        for row in worksheet.iter_rows(values_only=True):

            values = [
                str(value).strip() if value is not None else ""
                for value in row
            ]

            if any(values):
                rows.append(values)

        total_rows += len(rows)

        sheets.append(
            {
                "sheet_name": worksheet.title,
                "row_count": len(rows),
                "column_count": max(
                    (len(row) for row in rows),
                    default=0,
                ),
                "rows": rows,
            }
        )

    workbook.close()

    return {
        "text": "",
        "sheets": sheets,
        "sheet_count": len(sheets),
        "row_count": total_rows,
    }